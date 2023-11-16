import requests
import xlsxwriter
import time
from xlrd import open_workbook

old_repos=[]
new_repos=[]
base_url = "https://api.github.com/search/repositories"

def call_github(query, params,dates):

    params = {"q": "language:python+pushed:<2014-01-01", "per_page": 10}


    repositories = []
    count=0
    page_num = 1
    dates=dates
    while True:
        # time.sleep()
        params = {"q": "language:python+pushed:< {}".format(dates),
                  "page": page_num}
        response = requests.get("https://api.github.com/search/repositories?q=language:python+pushed:2016-06-01.."
                                "2017-01-01",
                                headers = {'Authorization': 'token %s' % "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" ,
           'Accept': 'application/vnd.github.v3+json'})
        page_num+=1
        if response.status_code == 200:
            data = response.json()
            print(len(repositories))
            repositories.extend(data["items"])
            # Check if there are more pages
            if "next" in response.links:
                print(count)
                count+=1
                # print(response.links)
                # # Update the 'page' parameter to get the next page
                # params["page"] = response.links["next"]["url"].split("page=")[1]
            else:
                break  # No more pages, exit the loop
        else:
            print(f"Error: {response.status_code}")
            print(response.text)

            break

    print(len(repositories))
    from datetime import datetime

    for j in repositories:
        date_str = j['pushed_at']

        # Parse the string into a datetime object
        date_object = datetime.strptime(date_str, '%Y-%m-%dT%H:%M:%SZ')

        # Extract the year
        year = date_object.year
        print(year)
        if year>2020:
            old_repos.append({"repo_name":j['name'], "owner":j['owner']['login'],"active_year":year})
    return old_repos

def call_github1(url, params):


    response = requests.get("https://api.github.com/search/repositories?q=language:python+pushed:{}".format(params))


    data = response.json()
    results=[]
    for k,v in data.items():
        if k == "items":
            results = v
            break
    for j in results:
        new_repos.append({"repo_name":j['name'], "owner":j['owner']['login']})

    print(len(new_repos))
    return new_repos
import openpyxl
count=0
result=call_github("a","","2014-01-01")
print(len(result))
print(result)
sheet_name = 'Old_Repositories'
# xlsx_filename = 'repositories.xlsx'
xlsx_filename = 'actively_updated_repositories.xlsx'
# Load the existing workbook
workbook = openpyxl.load_workbook(xlsx_filename)
try:
    worksheet = workbook[sheet_name]
except Exception as e:
    print(str(e))
    workbook.create_sheet(title=sheet_name)
    worksheet = workbook[sheet_name]
worksheet.cell(row=1, column=1, value='Owner')
worksheet.cell(row=1, column=2, value='Repo Name')
worksheet.cell(row=1,column=3, value='Active Year')

for repo in result:
    count=count+1
    print(repo)
    # worksheet.write('A{}'.format(count), repo['owner'])
    # worksheet.write('B{}'.format(count), repo['repo_name'])
    worksheet.cell(row=count, column=1, value=repo['owner'])
    worksheet.cell(row=count, column=2, value=repo['repo_name'])
    worksheet.cell(row=count, column=3, value=repo['active_year'])

# Finally, close the Excel file
# via the close() method.
workbook.save(xlsx_filename)
workbook.close()



